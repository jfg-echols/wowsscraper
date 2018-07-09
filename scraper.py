#! python3
import webbrowser
import sys
import bs4
import requests
import openpyxl
import os
import xlsxwriter
import re

rootSite = 'http://wiki.wargaming.net/en/World_of_Warships'


#shipclasses = ['DD', 'CA', 'BB', 'CV']
shipclasses = ['DD', 'CA', 'BB']
# shipclasses = ['CA', 'BB']
# skiplist = ['St. Louis']

for shipclass in shipclasses:

    if shipclass == 'DD':
        shipsite = 'http://wiki.wargaming.net/en/Ship:Destroyers'
    elif shipclass == 'CA':
        shipsite = 'http://wiki.wargaming.net/en/Ship:Cruisers'
    elif shipclass == 'BB':
        shipsite = 'http://wiki.wargaming.net/en/Ship:Battleships'
    elif shipclass == 'CV':
        shipsite = 'http://wiki.wargaming.net/en/Ship:Aircraft_Carriers'

    ##### creating the resultfile
 
    shipfilename = shipclass+'results.xlsx'
    relpath = os.path.abspath(os.path.dirname(__file__))
    resultfile = relpath+"\\"+shipfilename

    if os.path.exists(resultfile):
        os.remove(resultfile)
        print('deleted '+resultfile)
    print('creating file with xlswriter')
    xlswritership = xlsxwriter.Workbook(shipfilename)
    xlswritership.close()
    shipsheet = openpyxl.load_workbook(shipfilename)
    firstsheet = shipsheet['Sheet1']
    firstsheet.title = 'BattleLevel 1'
    shipsheet.create_sheet('BattleLevel 2')
    shipsheet.create_sheet('BattleLevel 3')
    shipsheet.create_sheet('BattleLevel 4')
    shipsheet.create_sheet('BattleLevel 5')
    shipsheet.create_sheet('BattleLevel 6')
    shipsheet.create_sheet('BattleLevel 7')
    shipsheet.create_sheet('BattleLevel 8')
    shipsheet.create_sheet('BattleLevel 9')
    shipsheet.create_sheet('BattleLevel 10')
    #resultfile created

    ##### setting the top row of each sheet

    toprow = [
        'shipName',
        'shipClass',
        'shipCountry',
        'shipTier',
        'shipCurrency',
        'shipCost',
        'shipHitPoints',
        'shipResearch',
        'shipMainGun',
        'shipMainGunCount',
        'shipMainGunROF',
        'shipMainGunReload',
        'shipMainGunRotate',
        'shipMainGun180',
        'shipMainGunRange',
        'shipMainGunDispersion',
        'shipMainGunHEShell',
        'shipMainGunHEDam',
        'shipMainGunHEFire',
        'shipMainGunHEVel',
        'shipMainGunHEWeight',
        'shipMainGunAPShell',
        'shipMainGunAPDam',
        'shipMainGunAPVel',
        'shipMainGunAPWeight'
    ]
    for sheet in shipsheet.worksheets:
        # print(toprow)
        for n in range(1,len(toprow)+1):
            sheet.cell(row=1,column=n).value = toprow[n-1]


    #testing the root site

    # rootres = requests.get(rootsite)
    # rootres.raise_for_status()
    # rootSoup = bs4.BeautifulSoup(rootres.text)

    ##TODO - turn this into something that isn't restricted to DD
    #get destroyer site
    pagereq = requests.get(shipsite)
    pagereq.raise_for_status()
    pagesoup = bs4.BeautifulSoup(pagereq.text, "html.parser")


    #print japanese DD names
    nationDivs = pagesoup.find_all("div",class_="wot-frame-1")

    for nation in nationDivs:
        
        print('====================')
        print(nation.find("h2").text)
        print('====================')
        
        
        nationShips = nation.find_all("div",class_="tleft")
        for thisShip in nationShips:
            fulllink = 'http://wiki.wargaming.net'+(thisShip.find_all("a"))[1].attrs['href']
            getshipname = (thisShip.find_all("a"))[1].text
            print('regex time')
            if (re.search("St.*Louis",getshipname)) or (re.search("Nueve.de.Julio",getshipname)):
                print('found it')
            else:
          
                
                shiparray = {}
                
                shiparray['shipName'] = getshipname



                print('----- ' + shiparray['shipName'])
                # print(fulllink)
                thisShipPage = requests.get(fulllink)
                thisShipPage.raise_for_status()
                thisShipSoup = bs4.BeautifulSoup(thisShipPage.text, "html.parser")

                

                #gets class, country, tier
                perfdiv = thisShipSoup.find('div',class_='b-performance_position')
                shiparray['shipClass'] = perfdiv.text.split(' | ')[0]
                shiparray['shipCountry'] = perfdiv.text.split(' | ')[1]
                shiparray['shipTier'] = perfdiv.text.split(' | ')[2]
                # print(shiparray)
                #gets general, armament, toobs, maneuverability, concealment
                shipstats = thisShipSoup.find('div',class_='b-performance_border').find_all('div',class_='gw-popup-card b-tech-nav_item__opened')
                for statgroup in shipstats:
                    groupName = statgroup.find('div',class_='b-performance_title gw-popup-card_head js-tech-nav_head').text
                    # print('--')
                    # print(groupName)
                    # print('--')
                    stats = statgroup.find('div',class_='gw-popup-card_content').find_all('tr')
                    
                    ###GENERAL STAT GROUP

                    if groupName == 'General':
                        
                        # shipCurrency = ''
                        # shipCost = ''
                        # shipHitPoints = ''
                        # shipResearch = ''
                        for stat in stats:
                            # print('--------------')
                            ##TODO - strip whitespace from variables                    
                            if stat.find('span',class_='t-performance_left').text == 'Purchase price':
                                shiparray['shipCurrency'] = stat.find('img')['alt']
                                shiparray['shipCost'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Hit Points':
                                shiparray['shipHitPoints'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Research price':
                                shiparray['shipResearch'] = stat.find('span',class_='t-performance_right').text
                            
                        # if shipCost != 'promo':
                        #     print('Currency: '+shipCurrency)
                        # print('Cost: '+shipCost)
                        # print('HP: '+shipHitPoints)
                        # if shipResearch != '':
                        #     print('ResearchXP: '+shipResearch)
                        #print(stats)
                    
                    ###MAIN BATTERY

                    elif groupName == 'Main Battery':
                        # shipMainGun = ''
                        # shipMainGunCount = ''
                        # shipMainGunROF = ''
                        # shipMainGunReload = ''
                        # shipMainGunRotate = ''
                        # shipMainGun180 = ''
                        # shipMainGunRange = ''
                        # shipMainGunDispersion = ''
                        # shipMainGunHEShell = ''
                        # shipMainGunHEDam = ''
                        # shipMainGunHEFire = ''
                        # shipMainGunHEVel = ''
                        # shipMainGunHEWeight = ''
                        # shipMainGunAPShell = ''
                        # shipMainGunAPDam = ''
                        # shipMainGunAPVel = ''
                        # shipMainGunAPWeight = ''

                        shiparray['shipMainGun'] = stats[0].find('span',class_='t-performance_left').text
                        shiparray['shipMainGunCount'] = stats[0].find('span',class_='t-performance_right').text

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Rate of Fire':
                                shiparray['shipMainGunROF'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Reload Time':
                                shiparray['shipMainGunReload'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Rotation Speed':
                                shiparray['shipMainGunRotate'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == '180 Degree Turn Time':
                                shiparray['shipMainGun180'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Firing Range':
                                shiparray['shipMainGunRange'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum Dispersion':
                                shiparray['shipMainGunDispersion'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell':
                                shiparray['shipMainGunHEShell'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum HE Shell Damage':
                                shiparray['shipMainGunHEDam'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Chance of Fire on Target Caused by HE Shell':
                                shiparray['shipMainGunHEFire'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial HE Shell Velocity':
                                shiparray['shipMainGunHEVel'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell Weight':
                                shiparray['shipMainGunHEWeight'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell':
                                shiparray['shipMainGunAPShell'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum AP Shell Damage':
                                shiparray['shipMainGunAPDam'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial AP Shell Velocity':
                                shiparray['shipMainGunAPVel'] = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shiparray['shipMainGunAPWeight'] = stat.find('span',class_='t-performance_right').text
                        
                        # print(shipMainGun)
                        # print(shipMainGunCount)
                        # print(shipMainGunROF)
                        # print(shipMainGunReload)
                        # print(shipMainGunRotate)
                        # print(shipMainGun180)
                        # print(shipMainGunRange)
                        # print(shipMainGunDispersion)
                        # print(shipMainGunHEShell)
                        # print(shipMainGunHEDam)
                        # print(shipMainGunHEFire)
                        # print(shipMainGunHEVel)
                        # print(shipMainGunHEWeight)
                        # print(shipMainGunAPShell)
                        # print(shipMainGunAPDam)
                        # print(shipMainGunAPVel)
                        # print(shipMainGunAPWeight)

                    ###Secondary Batteries

                    elif groupName == 'Secondary Armament #1':
                        shipSec1Gun = ''
                        shipSec1GunCount = ''
                        shipSec1GunROF = ''
                        shipSec1GunReload = ''
                        shipSec1GunRotate = ''
                        shipSec1Gun180 = ''
                        shipSec1GunRange = ''
                        shipSec1GunDispersion = ''
                        shipSec1GunHEShell = ''
                        shipSec1GunHEDam = ''
                        shipSec1GunHEFire = ''
                        shipSec1GunHEVel = ''
                        shipSec1GunHEWeight = ''
                        shipSec1GunAPShell = ''
                        shipSec1GunAPDam = ''
                        shipSec1GunAPVel = ''
                        shipSec1GunAPWeight = ''

                        shipSec1Gun = stats[0].find('span',class_='t-performance_left').text
                        shipSec1GunCount = stats[0].find('span',class_='t-performance_right').text

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Rate of Fire':
                                shipSec1GunROF = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Reload Time':
                                shipSec1GunReload = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Rotation Speed':
                                shipSec1GunRotate = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == '180 Degree Turn Time':
                                shipSec1Gun180 = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Firing Range':
                                shipSec1GunRange = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum Dispersion':
                                shipSec1GunDispersion = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell':
                                shipSec1GunHEShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum HE Shell Damage':
                                shipSec1GunHEDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Chance of Fire on Target Caused by HE Shell':
                                shipSec1GunHEFire = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial HE Shell Velocity':
                                shipSec1GunHEVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell Weight':
                                shipSec1GunHEWeight = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell':
                                shipSec1GunAPShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum AP Shell Damage':
                                shipSec1GunAPDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial AP Shell Velocity':
                                shipSec1GunAPVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shipSec1GunAPWeight = stat.find('span',class_='t-performance_right').text

                    elif groupName == 'Secondary Armament #2':
                        shipSec2Gun = ''
                        shipSec2GunCount = ''
                        shipSec2GunROF = ''
                        shipSec2GunReload = ''
                        shipSec2GunRotate = ''
                        shipSec2Gun180 = ''
                        shipSec2GunRange = ''
                        shipSec2GunDispersion = ''
                        shipSec2GunHEShell = ''
                        shipSec2GunHEDam = ''
                        shipSec2GunHEFire = ''
                        shipSec2GunHEVel = ''
                        shipSec2GunHEWeight = ''
                        shipSec2GunAPShell = ''
                        shipSec2GunAPDam = ''
                        shipSec2GunAPVel = ''
                        shipSec2GunAPWeight = ''

                        shipSec2Gun = stats[0].find('span',class_='t-performance_left').text
                        shipSec2GunCount = stats[0].find('span',class_='t-performance_right').text

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Rate of Fire':
                                shipSec2GunROF = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Reload Time':
                                shipSec2GunReload = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Rotation Speed':
                                shipSec2GunRotate = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == '180 Degree Turn Time':
                                shipSec2Gun180 = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Firing Range':
                                shipSec2GunRange = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum Dispersion':
                                shipSec2GunDispersion = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell':
                                shipSec2GunHEShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum HE Shell Damage':
                                shipSec2GunHEDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Chance of Fire on Target Caused by HE Shell':
                                shipSec2GunHEFire = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial HE Shell Velocity':
                                shipSec2GunHEVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell Weight':
                                shipSec2GunHEWeight = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell':
                                shipSec2GunAPShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum AP Shell Damage':
                                shipSec2GunAPDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial AP Shell Velocity':
                                shipSec2GunAPVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shipSec2GunAPWeight = stat.find('span',class_='t-performance_right').text

                    elif groupName == 'Secondary Armament #3':
                        shipSec3Gun = ''
                        shipSec3GunCount = ''
                        shipSec3GunROF = ''
                        shipSec3GunReload = ''
                        shipSec3GunRotate = ''
                        shipSec3Gun180 = ''
                        shipSec3GunRange = ''
                        shipSec3GunDispersion = ''
                        shipSec3GunHEShell = ''
                        shipSec3GunHEDam = ''
                        shipSec3GunHEFire = ''
                        shipSec3GunHEVel = ''
                        shipSec3GunHEWeight = ''
                        shipSec3GunAPShell = ''
                        shipSec3GunAPDam = ''
                        shipSec3GunAPVel = ''
                        shipSec3GunAPWeight = ''

                        shipSec3Gun = stats[0].find('span',class_='t-performance_left').text
                        shipSec3GunCount = stats[0].find('span',class_='t-performance_right').text

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Rate of Fire':
                                shipSec3GunROF = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Reload Time':
                                shipSec3GunReload = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Rotation Speed':
                                shipSec3GunRotate = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == '180 Degree Turn Time':
                                shipSec3Gun180 = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Firing Range':
                                shipSec3GunRange = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum Dispersion':
                                shipSec3GunDispersion = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell':
                                shipSec3GunHEShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum HE Shell Damage':
                                shipSec3GunHEDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Chance of Fire on Target Caused by HE Shell':
                                shipSec3GunHEFire = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial HE Shell Velocity':
                                shipSec3GunHEVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell Weight':
                                shipSec3GunHEWeight = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell':
                                shipSec3GunAPShell = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum AP Shell Damage':
                                shipSec3GunAPDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial AP Shell Velocity':
                                shipSec3GunAPVel = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shipSec3GunAPWeight = stat.find('span',class_='t-performance_right').text

                    ### Torpedo Tubes

                    elif groupName == 'Torpedo Tubes':
                        shipTorpTube = ''
                        shipTorpCount = ''
                        shipTorpReload = ''
                        shipTorpRorate = ''
                        shipTorp180 = ''
                        shipTorp = ''
                        shipTorpDam = ''
                        shipTorpSpeed = ''
                        shipTorpRange = ''

                        shipTorpTube = stats[0].find('span',class_='t-performance_left').text
                        shipTorpCount = stats[0].find('span',class_='t-performance_right').text

                        for stat in stats: 
                            if stat.find('span',class_='t-performance_left').text == 'Initial HE Shell Velocity':
                                shipTorpReload = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'HE Shell Weight':
                                shipTorpRorate = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell':
                                shipTorp180 = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Maximum AP Shell Damage':
                                shipTorp = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Initial AP Shell Velocity':
                                shipTorpDam = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shipTorpSpeed = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'AP Shell Weight':
                                shipTorpRange = stat.find('span',class_='t-performance_right').text


                    elif groupName == 'AA Defense':
                        ####FUCKING AA GUNS
                        #### for each ship's count of AA guns, add them to an array. When it's time to print, take the highest number of AA guns and make that the number of AA guns in the sheet
                        
                        ship0AA = ''
                        ship0AACount = ''
                        ship0AADps = ''
                        ship0AArange = ''
                        ship1AA = ''
                        ship1AACount = ''
                        ship1AADps = ''
                        ship1AArange = ''
                        ship2AA = ''
                        ship2AACount = ''
                        ship2AADps = ''
                        ship2AArange = ''
                        ship3AA = ''
                        ship3AACount = ''
                        ship3AADps = ''
                        ship3AArange = ''
                        ship4AA = ''
                        ship4AACount = ''
                        ship4AADps = ''
                        ship4AArange = ''
                        
                        class AAgun:
                            gun = ""
                            guncount = ""
                            gunrange = ""
                            gundps = ""

                            def __init__(self,gunname,count,grange,dps):
                                self.gun = gunname
                                self.guncount = count
                                self.gunrange = grange
                                self.gundps = dps

                            def calculatedps(self):
                                turrets = self.guncount.split('x')[0] 
                                barrelsper = self.guncount.split('x')[1]
                                totaldps = float(turrets)*float(barrelsper)*float(self.gundps)
                                return totaldps
                            def print (self):
                                print("this is my fun gun")
                                print(self.gun)
                                print(self.guncount)
                                print(self.gunrange)
                                print(self.gundps)

                        shipAAguns = []          
                        statcount = range(0,len(stats),1)
                        thegun = ""
                        thecount = ""
                        therange = ""
                        thedps = ""
                        for i in statcount:
                            # print(i)
                            if i%3 == 0:
                                # print(str(i)+"%3=0")
                                # print("turn this into the gun name and count")
                                thegun = stats[0].find('span',class_='t-performance_left').text
                                thecount = stats[0].find('span',class_='t-performance_right').text.replace("pcs.","").replace(' ','')
                            if i%3 ==1:
                                # print(str(i)+"%3=1")
                                # print("this is going to be the dps")
                                thedps = stats[1].find('span',class_='t-performance_right').text
                            if i%3 == 2:
                                # print(str(i)+"%3=2")
                                # print("this is going to be the range")
                                therange = stats[2].find('span',class_='t-performance_right').text
                                newgun = AAgun(thegun,thecount,thedps,therange)
                                shipAAguns.append(newgun)
                                thegun = ""
                                thecount = ""
                                thedps = ""
                                therange = ""
                        # for gun in shipAAguns:
                            # gun.print()

                            
                        
                        # for stat in stats:
                        #     print(stat)
                        #     # if line doesn't start with '...'


                    ##TODO - multiple AA Defense
                    
                    ### Maneuverability
                    
                    elif groupName == 'Maneuverability':
                        shipSpeed = ''
                        shipTurnRadius = ''
                        shipRudderShift = ''

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Maximum Speed':
                                shipSpeed = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Turning Circle Radius':
                                shipTurnRadius = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Rudder Shift Time':
                                shipRudderShift = stat.find('span',class_='t-performance_right').text

                    ### Concealment

                    elif groupName == 'Concealment':
                        shipSurfaceDetect = ''
                        shipAirDetect = ''

                        for stat in stats:
                            if stat.find('span',class_='t-performance_left').text == 'Surface Detectability Range':
                                shipSurfaceDetect = stat.find('span',class_='t-performance_right').text
                            if stat.find('span',class_='t-performance_left').text == 'Air Detectability Range':
                                shipAirDetect = stat.find('span',class_='t-performance_right').text

                shipBattleLevels = ""    
                for shipBattleLevel in thisShipSoup.find('span',class_='b-battles-levels_interval'):
                    if shipBattleLevels == "":
                        shipBattleLevels = shipBattleLevel.text
                    else:
                        shipBattleLevels = shipBattleLevels + "," + shipBattleLevel.text
                print("levels:"+shipBattleLevels)
                for level in shipBattleLevels.split(','):
                    # print('battle level '+level)
                    thesheet = shipsheet['BattleLevel '+level]
                    thisrow = thesheet.max_row+1
                    # print('next free row is '+str(thisrow))
                    for shipvalue in shiparray:
                        
                        # print('shipvalue we are writing is '+shipvalue)

                        for i in range(1,thesheet.max_column+1):
                            if thesheet.cell(row=1,column=i).value == shipvalue:
                                # print('the column for this value is '+str(i))
                                # print('saving ')
                                thesheet.cell(row=thisrow,column=i).value = shiparray[shipvalue]
                
                # input("Press a key to continue...")        
    shipsheet.save(shipfilename)